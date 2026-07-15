#!/usr/bin/env python3
"""Parse CUTLASS GEMM output and write results directly into the XLSX template."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on the target Linux environment
    load_workbook = None


CASE_RE = re.compile(
    r"^\[(?P<index>\d+)/(?P<total>\d+)\]\s+"
    r"L=(?P<context_lengths>\S+)\s+ops=(?P<source_operations>.+?)\s+"
    r"MxNxK=(?P<m>\d+)x(?P<n>\d+)x(?P<k>\d+)\s*$"
)
LEGACY_CASE_RE = re.compile(
    r"^\[(?P<index>\d+)/(?P<total>\d+)\]\s+"
    r"L=(?P<context_lengths>\d+)\s+op=(?P<source_operations>.+?)\s+"
    r"MxNxK=(?P<m>\d+)x(?P<n>\d+)x(?P<k>\d+)\s*$"
)
BEST_RE = re.compile(r"^Best configuration:\s*(?P<configuration>.+?)\s*$")
GFLOPS_RE = re.compile(r"^\s*gflops:\s*(?P<gflops>[0-9]+(?:\.[0-9]+)?)\s*$")

SHEET_NAME = "环境对比"
HEADER_ROW = 5
FIRST_DATA_ROW = 6


@dataclass(frozen=True)
class ParsedResult:
    context_lengths: str
    source_operations: str
    m: int
    n: int
    k: int
    best_configuration: str
    gflops: float

    @property
    def shape(self) -> tuple[int, int, int]:
        return self.m, self.n, self.k


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Parse CUTLASS gemm output and update the CUTLASS columns in the "
            "performance comparison XLSX workbook. HGEMM data is preserved."
        )
    )
    parser.add_argument("--log", required=True, type=Path)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("gemm_performance_comparison.xlsx"),
        help="XLSX template to read (default: gemm_performance_comparison.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output XLSX path; omit to update --workbook in place",
    )
    return parser.parse_args()


def parse_log(path: Path) -> list[ParsedResult]:
    current_case: dict[str, str] | None = None
    pending_configuration: str | None = None
    results: dict[tuple[int, int, int], ParsedResult] = {}

    with path.open("r", encoding="utf-8", errors="replace") as stream:
        for raw_line in stream:
            line = raw_line.rstrip("\r\n")
            case_match = CASE_RE.match(line) or LEGACY_CASE_RE.match(line)
            if case_match:
                current_case = case_match.groupdict()
                pending_configuration = None
                continue

            best_match = BEST_RE.match(line)
            if best_match and current_case:
                pending_configuration = best_match.group("configuration")
                continue

            gflops_match = GFLOPS_RE.match(line)
            if gflops_match and current_case and pending_configuration:
                result = ParsedResult(
                    context_lengths=current_case["context_lengths"],
                    source_operations=current_case["source_operations"],
                    m=int(current_case["m"]),
                    n=int(current_case["n"]),
                    k=int(current_case["k"]),
                    best_configuration=pending_configuration,
                    gflops=float(gflops_match.group("gflops")),
                )
                results[result.shape] = result
                pending_configuration = None

    if not results:
        raise ValueError(
            f"No complete cases found in {path}. The log must contain case headers "
            "and gemm's Best configuration/GFLOPS output."
        )
    return list(results.values())


def find_shape_rows(sheet) -> dict[tuple[int, int, int], int]:
    expected_headers = {
        "C": "M",
        "D": "N",
        "E": "K",
        "F": "CUTLASS最佳配置（脚本回填）",
        "G": "CUTLASS GFLOPS（脚本回填）",
    }
    for column, expected in expected_headers.items():
        actual = sheet[f"{column}{HEADER_ROW}"].value
        if actual != expected:
            raise ValueError(
                f"Unexpected template header {column}{HEADER_ROW}: {actual!r}; "
                f"expected {expected!r}."
            )

    rows: dict[tuple[int, int, int], int] = {}
    for row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in (3, 4, 5)]
        if all(isinstance(value, (int, float)) for value in values):
            shape = tuple(int(value) for value in values)
            if shape in rows:
                raise ValueError(f"Duplicate M/N/K in XLSX template: {shape}")
            rows[shape] = row
    return rows


def update_workbook(template: Path, output: Path, results: list[ParsedResult]) -> None:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to update XLSX files. Install it with: "
            "python3 -m pip install openpyxl"
        )
    if template.suffix.lower() != ".xlsx" or output.suffix.lower() != ".xlsx":
        raise ValueError("--workbook and --output must use the .xlsx extension")
    if not template.exists():
        raise FileNotFoundError(f"XLSX template does not exist: {template}")

    workbook = load_workbook(template)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    shape_rows = find_shape_rows(sheet)

    missing = [result.shape for result in results if result.shape not in shape_rows]
    if missing:
        formatted = ", ".join(f"{m}x{n}x{k}" for m, n, k in missing)
        raise ValueError(f"M/N/K not found in XLSX template: {formatted}")

    for result in results:
        row = shape_rows[result.shape]
        sheet.cell(row, 1).value = result.context_lengths
        sheet.cell(row, 2).value = result.source_operations
        sheet.cell(row, 6).value = result.best_configuration
        sheet.cell(row, 7).value = result.gflops

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    args = parse_args()
    output = args.output or args.workbook
    try:
        results = parse_log(args.log)
        update_workbook(args.workbook, output, results)
    except (OSError, RuntimeError, ValueError) as error:
        print(f"error: {error}", file=sys.stderr)
        return 1

    print(
        f"Updated {len(results)} CUTLASS cases in {output.resolve()}; "
        "HGEMM values, formulas, and formatting were preserved."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
